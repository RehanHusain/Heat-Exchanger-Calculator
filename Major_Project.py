import tkinter as tk
import numpy as np
from PIL import ImageTk, Image


# ------------ Module for Excel
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
wb = Workbook()

wb = load_workbook('MPExcel.xlsx')
ws = wb.active

# ---------------------
window = tk.Tk()
window.geometry("1150x500")
window.title("Application for Solving Parameters of Equipment Design")
window.iconbitmap("calcicon.ico")


# -----------------Creating Frames

def show_frame(frame):
    frame.tkraise()


frame0 = tk.Frame(window, bg='#3B3B3B')
frame2 = tk.Frame(window, bg='#3B3B3B')
frame3 = tk.Frame(window)
frame4 = tk.Frame(window)
frame5 = tk.Frame(window)
frame6 = tk.Frame(window)
frame7 = tk.Frame(window)
frame8 = tk.Frame(window, bg='#3B3B3B')
frame9 = tk.Frame(window)
frame10 = tk.Frame(window)
frame11 = tk.Frame(window)

for frame in (frame0, frame2, frame3, frame4, frame5, frame6, frame7, frame8, frame9, frame10, frame11):
    frame.grid(row=0, column=0, sticky="nsew")

# -----------Button On Frame 0 i.e Starting
f1btn = tk.Button(frame0, text="Next", font=("orbitron", 24, 'bold'), foreground="black", bg='white', padx=20, pady=1,
                  command=lambda: show_frame(frame2))
f1btn.place(x=815, y=377)

# -------------- Putting Image on frame 0
click = ImageTk.PhotoImage(Image.open('calculator.png'))
panel = tk.Label(frame0, image=click)
panel.place(x=700, y=20)

# ---------------- Frame 0 - Designing
frameing = tk.LabelFrame(frame0, padx=40, pady=40)
frameing.grid(row=1, column=1)
made1 = tk.Label(frameing, text="Designed By:-", font=("Times New Roman", 20, 'bold'), fg='red')
made1.grid(row=1, column=7)
name1 = tk.Label(frameing, text="Rehan Husain", font=("Times New Roman", 20, 'bold'))
name1.grid(row=2, column=7)


lb2 = tk.Label(frame0, text="\nApplication for Solving Parameters \nof Equipment Design\n ",
               font=("Times New Roman", 20, 'bold'))
lb2.grid(row=10, column=1)

# ----------- Frame 2
frameing2 = tk.LabelFrame(frame2, padx=50, pady=50)
frameing2.place(x=600, y=80)
frameing1 = tk.LabelFrame(frame2, padx=50, pady=50)
frameing1.place(x=200, y=80)

# ----------- Putting Back Button on Different Frames
f2btn = tk.Button(frame8, text="Back", padx=30, pady=10, command=lambda: show_frame(frame2))
f2btn.place(x=5, y=235)

f2btn = tk.Button(frame6, text="Back", padx=30, pady=10, command=lambda: show_frame(frame2))
f2btn.place(x=5, y=400)

f2btn = tk.Button(frame2, text="Back", padx=30, pady=10, command=lambda: show_frame(frame0))
f2btn.place(x=5, y=400)


# --------------- Frameing 1
lbl1 = tk.Label(frameing1, text="Shell and Tube\nHeat Exchanger", font=("Times New Roman", 20, 'bold'))
lbl1.grid(row=1, column=1)
f2btn = tk.Button(frameing1, text="Explore", font=("Arial Black", 10), fg='white', bg='#645394', padx=30, pady=10,
                  command=lambda: show_frame(frame7))
f2btn.grid(row=2, column=1)

# --------------- Frameing 2
lbl1 = tk.Label(frameing2, text="Shell and Tube\nHeat Exchanger\n(Excel Optimization)",
                font=("Times New Roman", 20, 'bold'))
lbl1.grid(row=1, column=1)
f2btn = tk.Button(frameing2, text="Explore", font=("Arial Black", 10), fg='white', bg='orange', padx=30, pady=10,
                  command=lambda: show_frame(frame6))
f2btn.grid(row=2, column=1)

# ------------------------------shell and tube heat exchanger ------------------------------------------------

clickpitch = tk.StringVar()

clicktrsq = ["Triangular pitch", "Square pitch"]

# ----------- Default as "Triangular pitch"
clickpitch.set(clicktrsq[0])


# -------------------------- Re vs Jh

# -----------------Setting Jh anf Jf For Different BS and L/di
def heatfac():
    clickheat = tk.StringVar()
    optionheat = ["24", "48", "120", "240", "500"]
    clickheat.set(optionheat[0])

    clickheat1 = tk.StringVar()
    optionheat1 = ["15", "25", "35", "45"]
    clickheat1.set(optionheat1[0])

    # ------- Pitch for K and N values
    pitch = tk.StringVar()

    trsq = ["Triangular pitch", "Square pitch"]

    pitch.set(trsq[0])
    passes = tk.StringVar()

    noof = ["1", "2", "4", "6", "8"]

    passes.set(noof[0])

    # --------------------- Tubeside Jh
    # ------- Using Graph of Re and Jh For Diff L/Di
    def finding1():
        if clickheat.get() == "24":
            z = float(x1.get())
            if z <= 2000:
                yo = (-0.131 * (np.log(z))) + 1.00024
                y1.delete(0, len(str(float(yo))))
                y1.insert(0, str(yo))
            elif z > 2000 and z <= 10000:
                yo = (-6.25 * (10 ** (-8)) * z) + 0.004625
                y1.delete(0, len(str(float(yo))))
                y1.insert(0, str(yo))
            elif z > 10000 and z <= 1000000:
                yo = (-2.22 * (10 ** (-9)) * z) + 0.004022
                y1.delete(0, len(str(float(yo))))
                y1.insert(0, str(yo))

        elif clickheat.get() == "48":
            z = float(x1.get())
            if z <= 2000:
                yo = (-4.83417 * (10 ** (-5)) * z) + 0.1346
                y1.delete(0, len(str(float(yo))))
                y1.insert(0, str(yo))
            elif z > 2000 and z <= 10000:
                yo = (0.000000025 * z) + 0.00375
                y1.delete(0, len(str(float(yo))))
                y1.insert(0, str(yo))
            elif z > 10000 and z <= 1000000:
                yo = (-2.22 * (10 ** (-9)) * z) + 0.004022
                y1.delete(0, len(str(float(yo))))
                y1.insert(0, str(yo))

        elif clickheat.get() == "120":
            z = float(x1.get())
            if z <= 2000:
                yo = (-3.87437 * (10 ** (-5)) * z) + 0.0803874
                y1.delete(0, len(str(float(yo))))
                y1.insert(0, str(yo))
            elif z > 2000 and z <= 10000:
                yo = (-7 * (10 ** (-11)) * z * z) + (9 * (10 ** (-7)) * z) + 0.0013
                y1.delete(0, len(str(float(yo))))
                y1.insert(0, str(yo))
            elif z > 10000 and z <= 1000000:
                yo = (-2.22 * (10 ** (-9)) * z) + 0.004022
                y1.delete(0, len(str(float(yo))))
                y1.insert(0, str(yo))

        elif clickheat.get() == "240":
            z = float(x1.get())
            if z <= 2000:
                yo = (-2.91457 * (10 ** (-5)) * z) + 0.06
                y1.delete(0, len(str(float(yo))))
                y1.insert(0, str(yo))
            elif z > 2000 and z <= 10000:
                yo = (-1 * (10 ** (-10)) * z * z) + (2 * (10 ** (-6)) * z) - 0.0006
                y1.delete(0, len(str(float(yo))))
                y1.insert(0, str(yo))
            elif z > 10000 and z <= 1000000:
                yo = (-2.22 * (10 ** (-9)) * z) + 0.004022
                y1.delete(0, len(str(float(yo))))
                y1.insert(0, str(yo))

        elif clickheat.get() == "500":
            z = float(x1.get())
            if z <= 2000:
                yo = (-2.13065 * (10 ** (-5)) * z) + 0.04421
                y1.delete(0, len(str(float(yo))))
                y1.insert(0, str(yo))
            elif z > 2000 and z <= 10000:
                yo = (-9 * (10 ** (-11)) * z * z) + (1 * (10 ** (-6)) * z) - 0.0007
                y1.delete(0, len(str(float(yo))))
                y1.insert(0, str(yo))
            elif z > 10000 and z <= 1000000:
                yo = (-2.22 * (10 ** (-9)) * z) + 0.004022
                y1.delete(0, len(str(float(yo))))
                y1.insert(0, str(yo))

    # --------------Tubeside Jf
    def finding2():
        z = float(x2.get())
        if z <= 2000:
            yo = (-0.0004 * z) + 0.804
            y2.delete(0, len(str(float(yo))))
            y2.insert(0, str(yo))
        elif z > 2000 and z <= 3200:
            yo = (-1 * (10 ** (-9)) * z * z) + (8 * (10 ** (-6)) * z) - 0.0075
            y2.delete(0, len(str(float(yo))))
            y2.insert(0, str(yo))
        elif z > 3200 and z <= 1000000:
            yo = (-4.3138 * (10 ** (-9)) * z) + 0.0061138
            y2.delete(0, len(str(float(yo))))
            y2.insert(0, str(yo))

    # Shell Side Jh
    def finding3():
        if clickheat1.get() == "15":
            z = float(x3.get())
            if z <= 140:
                yo = (2 * (10 ** (-5)) * (z ** 2)) + ((-0.0039) * z) + 0.2513
                y3.delete(0, len(str(float(yo))))
                y3.insert(0, str(yo))
            elif z > 140 and z <= 10000:
                yo = (7 * (10 ** (-10)) * (z ** 2)) + (-1 * (10 ** (-5)) * z) + 0.0303
                y3.delete(0, len(str(float(yo))))
                y3.insert(0, str(yo))
            elif z > 10000 and z <= 100000:
                yo = (1 * (10 ** (-12)) * (z ** 2)) + (-2 * (10 ** (-7)) * z) + 0.008
                y3.delete(0, len(str(float(yo))))
                y3.insert(0, str(yo))
            elif z > 100000 and z <= 1000000:
                yo = (-1.4444 * (10 ** (-9)) * z) + 0.002704
                y3.delete(0, len(str(float(yo))))
                y3.insert(0, str(yo))

        elif clickheat1.get() == "25":
            z = float(x3.get())
            if z <= 140:
                yo = (1 * (10 ** (-5)) * (z ** 2)) + ((-0.0032) * z) + 0.2257
                y3.delete(0, len(str(float(yo))))
                y3.insert(0, str(yo))
            elif z > 140 and z <= 10000:
                yo = (4 * (10 ** (-9)) * (z ** 2)) + (-4 * (10 ** (-5)) * z) + 0.0464
                y3.delete(0, len(str(float(yo))))
                y3.insert(0, str(yo))
            elif z > 10000 and z <= 100000:
                yo = (7 * (10 ** (-13)) * (z ** 2)) + (-1 * (10 ** (-7)) * z) + 0.0069
                y3.delete(0, len(str(float(yo))))
                y3.insert(0, str(yo))
            elif z > 100000 and z <= 1000000:
                yo = (-1.44444 * (10 ** (-9)) * z) + 0.0021444
                y3.delete(0, len(str(float(yo))))
                y3.insert(0, str(yo))

        elif clickheat1.get() == "35":
            z = float(x3.get())
            if z <= 140:
                yo = (1 * (10 ** (-5)) * (z ** 2)) + ((-0.0029) * z) + 0.1863
                y3.delete(0, len(str(float(yo))))
                y3.insert(0, str(yo))
            elif z > 140 and z <= 10000:
                yo = (9 * (10 ** (-10)) * (z ** 2)) + (-1 * (10 ** (-5)) * z) + 0.0351
                y3.delete(0, len(str(float(yo))))
                y3.insert(0, str(yo))
            elif z > 10000 and z <= 100000:
                yo = (9 * (10 ** (-13)) * (z ** 2)) + (-1 * (10 ** (-7)) * z) + 0.0063
                y3.delete(0, len(str(float(yo))))
                y3.insert(0, str(yo))
            elif z > 100000 and z <= 1000000:
                yo = (-1.4444 * (10 ** (-9)) * z) + 0.0020444
                y3.delete(0, len(str(float(yo))))
                y3.insert(0, str(yo))

        elif clickheat1.get() == "45":
            z = float(x3.get())
            if z <= 140:
                yo = (1 * (10 ** (-5)) * (z ** 2)) + ((-0.0031) * z) + 0.1874
                y3.delete(0, len(str(float(yo))))
                y3.insert(0, str(yo))
            elif z > 140 and z <= 10000:
                yo = (7 * (10 ** (-10)) * (z ** 2)) + (-1 * (10 ** (-5)) * z) + 0.032
                y3.delete(0, len(str(float(yo))))
                y3.insert(0, str(yo))
            elif z > 10000 and z <= 100000:
                yo = (5 * (10 ** (-13)) * (z ** 2)) + (-8 * (10 ** (-8)) * z) + 0.0058
                y3.delete(0, len(str(float(yo))))
                y3.insert(0, str(yo))
            elif z > 100000 and z <= 1000000:
                yo = (-1.4444 * (10 ** (-9)) * z) + 0.0019422
                y3.delete(0, len(str(float(yo))))
                y3.insert(0, str(yo))

    # Shell Side Jf
    def finding4():
        if clickheat1.get() == "15":
            z = float(x4.get())
            if z >= 10 and z <= 300:
                yo = (0.0002 * (z ** 2)) + (-0.074 * z) + 4.3203
                y4.delete(0, len(str(float(yo))))
                y4.insert(0, str(yo))
            elif z > 300 and z <= 4000:
                yo = (1 * (10 ** (-8)) * (z ** 2)) + (-6 * (10 ** (-5)) * z) + 0.1609
                y4.delete(0, len(str(float(yo))))
                y4.insert(0, str(yo))
            elif z > 4000 and z <= 100000:
                yo = (1 * (10 ** (-11)) * (z ** 2)) + (-1 * (10 ** (-6)) * z) + 0.0844
                y4.delete(0, len(str(float(yo))))
                y4.insert(0, str(yo))
            elif z > 100000 and z <= 1000000:
                yo = (2 * (10 ** (-14)) * (z ** 2)) + (-4 * (10 ** (-8)) * z) + 0.0509
                y4.delete(0, len(str(float(yo))))
                y4.insert(0, str(yo))

        elif clickheat1.get() == "25":
            z = float(x4.get())
            if z >= 10 and z <= 300:
                yo = (9 * (10 ** (-5)) * (z ** 2)) + (-0.0349 * z) + 2.2396
                y4.delete(0, len(str(float(yo))))
                y4.insert(0, str(yo))
            elif z > 300 and z <= 4000:
                yo = (1 * (10 ** (-8)) * (z ** 2)) + (-6 * (10 ** (-5)) * z) + 0.114
                y4.delete(0, len(str(float(yo))))
                y4.insert(0, str(yo))
            elif z > 4000 and z <= 100000:
                yo = (4 * (10 ** (-12)) * (z ** 2)) + (-7 * (10 ** (-7)) * z) + 0.0592
                y4.delete(0, len(str(float(yo))))
                y4.insert(0, str(yo))
            elif z > 100000 and z <= 1000000:
                yo = (3 * (10 ** (-14)) * (z ** 2)) + (-4 * (10 ** (-8)) * z) + 0.0391
                y4.delete(0, len(str(float(yo))))
                y4.insert(0, str(yo))

        elif clickheat1.get() == "35":
            z = float(x4.get())
            if z >= 10 and z <= 300:
                yo = (0.0001 * (z ** 2)) + (-0.048 * z) + 2.4668
                y4.delete(0, len(str(float(yo))))
                y4.insert(0, str(yo))
            elif z > 300 and z <= 4000:
                yo = (6 * (10 ** (-8)) * (z ** 2)) + (-0.0001 * z) + 0.1216
                y4.delete(0, len(str(float(yo))))
                y4.insert(0, str(yo))
            elif z > 4000 and z <= 100000:
                yo = (3 * (10 ** (-12)) * (z ** 2)) + (-6 * (10 ** (-7)) * z) + 0.0508
                y4.delete(0, len(str(float(yo))))
                y4.insert(0, str(yo))
            elif z > 100000 and z <= 1000000:
                yo = (2 * (10 ** (-14)) * (z ** 2)) + (-3 * (10 ** (-8)) * z) + 0.0312
                y4.delete(0, len(str(float(yo))))
                y4.insert(0, str(yo))

        elif clickheat1.get() == "45":
            z = float(x4.get())
            if z >= 10 and z <= 300:
                yo = (5 * (10 ** (-5)) * (z ** 2)) + (-0.0195 * z) + 1.6924
                y4.delete(0, len(str(float(yo))))
                y4.insert(0, str(yo))
            elif z > 300 and z <= 4000:
                yo = (7 * (10 ** (-9)) * (z ** 2)) + (-4 * (10 ** (-5)) * z) + 0.0787
                y4.delete(0, len(str(float(yo))))
                y4.insert(0, str(yo))
            elif z > 4000 and z <= 100000:
                yo = (2 * (10 ** (-12)) * (z ** 2)) + (-4 * (10 ** (-7)) * z) + 0.0398
                y4.delete(0, len(str(float(yo))))
                y4.insert(0, str(yo))
            elif z > 100000 and z <= 1000000:
                yo = (2 * (10 ** (-14)) * (z ** 2)) + (-3 * (10 ** (-8)) * z) + 0.0261
                y4.delete(0, len(str(float(yo))))
                y4.insert(0, str(yo))

    def searched():
        if pitch.get() == "Triangular pitch" and passes.get() == "1":
            u = 0.319
            v = 2.142
            y5.delete(0, len(str(float(u))))
            y5.insert(0, str(u))
            y6.delete(0, len(str(float(v))))
            y6.insert(0, str(v))
        elif pitch.get() == "Triangular pitch" and passes.get() == "2":
            u = 0.249
            v = 2.207
            y5.delete(0, len(str(float(u))))
            y5.insert(0, str(u))
            y6.delete(0, len(str(float(v))))
            y6.insert(0, str(v))
        elif pitch.get() == "Triangular pitch" and passes.get() == "4":
            u = 0.175
            v = 2.285
            y5.delete(0, len(str(float(u))))
            y5.insert(0, str(u))
            y6.delete(0, len(str(float(v))))
            y6.insert(0, str(v))
        elif pitch.get() == "Triangular pitch" and passes.get() == "6":
            u = 0.0743
            v = 2.499
            y5.delete(0, len(str(float(u))))
            y5.insert(0, str(u))
            y6.delete(0, len(str(float(v))))
            y6.insert(0, str(v))
        elif pitch.get() == "Triangular pitch" and passes.get() == "8":
            u = 0.0365
            v = 2.675
            y5.delete(0, len(str(float(u))))
            y5.insert(0, str(u))
            y6.delete(0, len(str(float(v))))
            y6.insert(0, str(v))

        elif pitch.get() == "Square pitch" and passes.get() == "1":
            u = 0.215
            v = 2.207
            y5.delete(0, len(str(float(u))))
            y5.insert(0, str(u))
            y6.delete(0, len(str(float(v))))
            y6.insert(0, str(v))
        elif pitch.get() == "Square pitch" and passes.get() == "2":
            u = 0.156
            v = 2.291
            y5.delete(0, len(str(float(u))))
            y5.insert(0, str(u))
            y6.delete(0, len(str(float(v))))
            y6.insert(0, str(v))
        elif pitch.get() == "Square pitch" and passes.get() == "4":
            u = 0.158
            v = 2.263
            y5.delete(0, len(str(float(u))))
            y5.insert(0, str(u))
            y6.delete(0, len(str(float(v))))
            y6.insert(0, str(v))
        elif pitch.get() == "Square pitch" and passes.get() == "6":
            u = 0.0402
            v = 2.617
            y5.delete(0, len(str(float(u))))
            y5.insert(0, str(u))
            y6.delete(0, len(str(float(v))))
            y6.insert(0, str(v))
        elif pitch.get() == "Square pitch" and passes.get() == "8":
            u = 0.0331
            v = 2.643
            y5.delete(0, len(str(float(u))))
            y5.insert(0, str(u))
            y6.delete(0, len(str(float(v))))
            y6.insert(0, str(v))

    tab = tk.Toplevel()
    tab.title("calculate points")
    tab.geometry("400x500")
    x = tk.Label(tab, text="For Tube Side", font=("Helvetica", 10))
    x.place(x=30, y=10)
    compoundheat = tk.OptionMenu(tab, clickheat, *optionheat)
    compoundheat.place(x=120, y=30)
    x = tk.Label(tab, text="L/D ratio", font=("Helvetica", 10))
    x.place(x=50, y=30)
    x = tk.Label(tab, text="Re", font=("Helvetica", 10))
    x.place(x=30, y=60)
    x1 = tk.Entry(tab, width=10, borderwidth=5)
    x1.place(x=50, y=60)
    y = tk.Label(tab, text="Jh", font=("Helvetica", 10))
    y.place(x=120, y=60)
    y1 = tk.Entry(tab, width=10, borderwidth=5)
    y1.place(x=140, y=60)
    btc11 = tk.Button(tab, text="convert", padx=10, pady=5, command=finding1)
    btc11.place(x=220, y=60)

    x = tk.Label(tab, text="Re", font=("Helvetica", 10))
    x.place(x=30, y=110)
    x2 = tk.Entry(tab, width=10, borderwidth=5)
    x2.place(x=50, y=110)
    y = tk.Label(tab, text="Jf", font=("Helvetica", 10))
    y.place(x=120, y=110)
    y2 = tk.Entry(tab, width=10, borderwidth=5)
    y2.place(x=140, y=110)
    btc12 = tk.Button(tab, text="convert", padx=10, pady=5, command=finding2)
    btc12.place(x=220, y=110)

    # Shell Side
    x = tk.Label(tab, text="For Shell Side", font=("Helvetica", 10))
    x.place(x=30, y=140)
    compoundheat1 = tk.OptionMenu(tab, clickheat1, *optionheat1)
    compoundheat1.place(x=130, y=170)
    x = tk.Label(tab, text="Baffle cut %", font=("Helvetica", 10))
    x.place(x=50, y=170)
    x = tk.Label(tab, text="Re", font=("Helvetica", 10))
    x.place(x=30, y=210)
    x3 = tk.Entry(tab, width=10, borderwidth=5)
    x3.place(x=50, y=210)
    y = tk.Label(tab, text="Jh", font=("Helvetica", 10))
    y.place(x=120, y=210)
    y3 = tk.Entry(tab, width=10, borderwidth=5)
    y3.place(x=140, y=210)
    btc11 = tk.Button(tab, text="convert", padx=10, pady=5, command=finding3)
    btc11.place(x=220, y=210)

    x = tk.Label(tab, text="Re", font=("Helvetica", 10))
    x.place(x=30, y=260)
    x4 = tk.Entry(tab, width=10, borderwidth=5)
    x4.place(x=50, y=260)
    y = tk.Label(tab, text="Jf", font=("Helvetica", 10))
    y.place(x=120, y=260)
    y4 = tk.Entry(tab, width=10, borderwidth=5)
    y4.place(x=140, y=260)
    btc11 = tk.Button(tab, text="convert", padx=10, pady=5, command=finding4)
    btc11.place(x=220, y=260)

    x = tk.Label(tab, text="Shell side numerals", font=("Helvetica", 10))
    x.place(x=30, y=290)
    x = tk.Label(tab, text="Pitch Type", font=("Helvetica", 10))
    x.place(x=50, y=320)
    set1 = tk.OptionMenu(tab, pitch, *trsq)
    set1.place(x=200, y=320)
    x = tk.Label(tab, text="No.of passes", font=("Helvetica", 10))
    x.place(x=50, y=360)
    set2 = tk.OptionMenu(tab, passes, *noof)
    set2.place(x=200, y=360)
    x = tk.Label(tab, text="K1", font=("Helvetica", 10))
    x.place(x=50, y=410)
    y5 = tk.Entry(tab, width=10, borderwidth=5)
    y5.place(x=80, y=410)
    x = tk.Label(tab, text="n1", font=("Helvetica", 10))
    x.place(x=50, y=450)
    y6 = tk.Entry(tab, width=10, borderwidth=5)
    y6.place(x=80, y=450)
    btc11 = tk.Button(tab, text="Search", padx=10, pady=5, command=searched)
    btc11.place(x=280, y=410)


# ---------------- Function for Calculating Value of Shell And Tube side
def shellatube():
    tabst = tk.Toplevel()
    tabst.title("Shell and Tube Heat Exchanger Output")
    tabst.geometry("1200x500")

    frameingtube = tk.LabelFrame(tabst, padx=30, pady=30)
    frameingtube.grid(row=1, column=1)
    frameingshell = tk.LabelFrame(tabst, padx=30, pady=30)
    frameingshell.grid(row=1, column=3)
    frameingoverall = tk.LabelFrame(tabst, padx=30, pady=30)
    frameingoverall.grid(row=2, column=2)

    # ----------------------------- Final TubeSIde Calculation
    mflow1 = (float(mflow2.get()) * float(spheat2.get()) * (float(inshell.get()) - float(outshell.get()))) / (
            (float(outtube.get()) - float(intube.get())) * float(spheat1.get()))
    heatflow1 = mflow1 * float(spheat1.get()) * (float(outtube.get()) - float(intube.get()))
    delt = ((float(inshell.get()) - float(outtube.get())) - (float(outshell.get()) - float(intube.get()))) / (
        np.log((float(inshell.get()) - float(outtube.get())) / (float(outshell.get()) - float(intube.get()))))
    areatube = heatflow1 / (float(uass.get()) * float(fowling.get()) * delt * 3600)
    area1 = (22 * (float(ind1.get()) ** 2)) / 28
    nooftube = (areatube * 7) / (22 * float(outd1.get()) * float(length.get()))
    velocity1 = (mflow1 * float(nofp.get())) / (nooftube * float(den1.get()) * area1 * 3600)
    reynold1 = (float(ind1.get()) * velocity1 * float(den1.get())) / float(vis1.get())
    pr1 = (float(vis1.get()) * float(spheat1.get())) / float(tcon1.get())

    # Nu1 = (h1 * float(ind1.get())) / (float(tcon1.get()))
    # h1 = (float(tcon1.get()) * reynold1 * (pr1 ** 0.33) * float(jht.get())) / float(ind1.get())

    Nu1 = reynold1 * float(jht.get()) * (pr1 ** 0.33)
    h1 = (Nu1 * float(tcon1.get())) / float(ind1.get())
    pressure1 = float(nofp.get()) * (((8 * float(jft.get()) * (float(length.get()) / float(ind1.get()))) + 2.5) * (float(den1.get()) * (velocity1 ** 2))) / 2000

    # tube
    lb2 = tk.Label(frameingtube, text="Tube Side", font=("Helvetica", 20))
    lb2.grid(row=1, column=2)
    lb2 = tk.Label(frameingtube, text="Mass flow rate", font=("Helvetica", 10))
    lb2.grid(row=2, column=1)
    lb2 = tk.Label(frameingtube, text=str(mflow1), font=("Helvetica", 10))
    lb2.grid(row=2, column=2)
    lb2 = tk.Label(frameingtube, text="Kg/hr", font=("Helvetica", 10))
    lb2.grid(row=2, column=3)
    lb2 = tk.Label(frameingtube, text="Heat transfer(Q)", font=("Helvetica", 10))
    lb2.grid(row=10, column=1)
    lb2 = tk.Label(frameingtube, text=str(heatflow1), font=("Helvetica", 10))
    lb2.grid(row=10, column=2)
    lb2 = tk.Label(frameingtube, text="J/hr", font=("Helvetica", 10))
    lb2.grid(row=10, column=3)
    lb2 = tk.Label(frameingtube, text="Area of tube", font=("Helvetica", 10))
    lb2.grid(row=3, column=1)
    lb2 = tk.Label(frameingtube, text=str(area1), font=("Helvetica", 10))
    lb2.grid(row=3, column=2)
    lb2 = tk.Label(frameingtube, text="m2", font=("Helvetica", 10))
    lb2.grid(row=3, column=3)
    lb2 = tk.Label(frameingtube, text="No. of tubes", font=("Helvetica", 10))
    lb2.grid(row=4, column=1)
    lb2 = tk.Label(frameingtube, text=str(nooftube), font=("Helvetica", 10))
    lb2.grid(row=4, column=2)
    lb2 = tk.Label(frameingtube, text="velocity in tube", font=("Helvetica", 10))
    lb2.grid(row=5, column=1)
    lb2 = tk.Label(frameingtube, text=str(velocity1), font=("Helvetica", 10))
    lb2.grid(row=5, column=2)
    lb2 = tk.Label(frameingtube, text="m/s", font=("Helvetica", 10))
    lb2.grid(row=5, column=3)
    lb2 = tk.Label(frameingtube, text="Reynolds no.", font=("Helvetica", 10))
    lb2.grid(row=6, column=1)
    lb2 = tk.Label(frameingtube, text=str(reynold1), font=("Helvetica", 10))
    lb2.grid(row=6, column=2)
    lb2 = tk.Label(frameingtube, text="Prandtl no.", font=("Helvetica", 10))
    lb2.grid(row=7, column=1)
    lb2 = tk.Label(frameingtube, text=str(pr1), font=("Helvetica", 10))
    lb2.grid(row=7, column=2)
    lb2 = tk.Label(frameingtube, text="Heat transfer coeff ", font=("Helvetica", 10))
    lb2.grid(row=8, column=1)
    lb2 = tk.Label(frameingtube, text=str(h1), font=("Helvetica", 10))
    lb2.grid(row=8, column=2)
    lb2 = tk.Label(frameingtube, text="W/m2C", font=("Helvetica", 10))
    lb2.grid(row=8, column=3)
    lb2 = tk.Label(frameingtube, text="Pressure", font=("Helvetica", 10))
    lb2.grid(row=9, column=1)
    lb2 = tk.Label(frameingtube, text=str(pressure1), font=("Helvetica", 10))
    lb2.grid(row=9, column=2)
    lb2 = tk.Label(frameingtube, text="kN/m2", font=("Helvetica", 10))
    lb2.grid(row=9, column=3)
    lb2 = tk.Label(frameingtube, text="Nusset No.", font=("Helvetica", 10))
    lb2.grid(row=11, column=1)
    lb2 = tk.Label(frameingtube, text=str(Nu1), font=("Helvetica", 10))
    lb2.grid(row=11, column=2)

    # ------------------------------Final ShellSide Calculation
    heatflow2 = float(mflow2.get()) * float(spheat2.get()) * (float(inshell.get()) - float(outshell.get()))
    bundledia = float(outd1.get()) * ((nooftube / float(k1s.get())) ** (1 / float(n1s.get())))
    pitch = 1.25 * float(outd1.get())
    shelldia = 1.1 * bundledia
    bafflespace = shelldia / 5
    area2 = ((pitch - float(outd1.get())) * shelldia * bafflespace) / pitch
    eqdia = 0
    if clickpitch.get() == "Triangular pitch":
        eqdia = 1.1 * ((pitch ** 2) - (0.917 * (float(outd1.get()) ** 2))) / float(outd1.get())
    elif clickpitch.get() == "Square pitch":
        eqdia = 1.27 * ((pitch ** 2) - (0.785 * (float(outd1.get()) ** 2))) / float(outd1.get())
    velocity2 = float(mflow2.get()) / (float(den2.get()) * area2 * 3600)
    reynold2 = (eqdia * velocity2 * float(den2.get())) / (float(vis2.get()))
    pr2 = float(vis2.get()) * float(spheat2.get()) / (float(tcon2.get()))
    Nu2 = reynold2 * float(jhs.get()) * (pr2 ** 0.33)
    h2 = (Nu2 * float(tcon2.get())) / eqdia
    pressure2 = (8 * float(jfs.get()) * shelldia * float(length.get()) * float(den2.get()) * (velocity2 ** 2)) / (
            eqdia * bafflespace * 2000)

    # ------------------ OverAll Heat Tranfer Coefficient
    Ucal = (((1 / h1) * (float(outd1.get()) / float(ind1.get()))) + (1 / h2) + (
            (float(outd1.get()) * np.log(float(outd1.get()) / float(ind1.get()))) / (
            2 * float(alloy.get()))) + float(fouls.get()) + (
                    float(foult.get()) * (float(outd1.get()) / float(ind1.get())))) ** (-1)
    error = ((Ucal - float(uass.get())) / Ucal) * 100

    # ------------------------------Final ShellSide Calculation
    lb2 = tk.Label(frameingshell, text="Shell Side", font=("Helvetica", 20))
    lb2.grid(row=1, column=2)
    lb2 = tk.Label(frameingshell, text="Heat transfer(Q)", font=("Helvetica", 10))
    lb2.grid(row=10, column=1)
    lb2 = tk.Label(frameingshell, text=str(heatflow2), font=("Helvetica", 10))
    lb2.grid(row=10, column=2)
    lb2 = tk.Label(frameingshell, text="J/hr", font=("Helvetica", 10))
    lb2.grid(row=10, column=3)
    lb2 = tk.Label(frameingshell, text="Shell diameter", font=("Helvetica", 10))
    lb2.grid(row=3, column=1)
    lb2 = tk.Label(frameingshell, text=str(shelldia), font=("Helvetica", 10))
    lb2.grid(row=3, column=2)
    lb2 = tk.Label(frameingshell, text="m", font=("Helvetica", 10))
    lb2.grid(row=3, column=3)
    lb2 = tk.Label(frameingshell, text="Baffle spacing", font=("Helvetica", 10))
    lb2.grid(row=3, column=1)
    lb2 = tk.Label(frameingshell, text=str(bafflespace), font=("Helvetica", 10))
    lb2.grid(row=3, column=2)
    lb2 = tk.Label(frameingshell, text="m", font=("Helvetica", 10))
    lb2.grid(row=3, column=3)
    lb2 = tk.Label(frameingshell, text="Equivalent diameter", font=("Helvetica", 10))
    lb2.grid(row=4, column=1)
    lb2 = tk.Label(frameingshell, text=str(eqdia), font=("Helvetica", 10))
    lb2.grid(row=4, column=2)
    lb2 = tk.Label(frameingshell, text="m", font=("Helvetica", 10))
    lb2.grid(row=4, column=3)
    lb2 = tk.Label(frameingshell, text="velocity", font=("Helvetica", 10))
    lb2.grid(row=5, column=1)
    lb2 = tk.Label(frameingshell, text=str(velocity2), font=("Helvetica", 10))
    lb2.grid(row=5, column=2)
    lb2 = tk.Label(frameingshell, text="m/s", font=("Helvetica", 10))
    lb2.grid(row=5, column=3)
    lb2 = tk.Label(frameingshell, text="Reynold no.", font=("Helvetica", 10))
    lb2.grid(row=6, column=1)
    lb2 = tk.Label(frameingshell, text=str(reynold2), font=("Helvetica", 10))
    lb2.grid(row=6, column=2)
    lb2 = tk.Label(frameingshell, text="Prandtl no.", font=("Helvetica", 10))
    lb2.grid(row=7, column=1)
    lb2 = tk.Label(frameingshell, text=str(pr2), font=("Helvetica", 10))
    lb2.grid(row=7, column=2)
    lb2 = tk.Label(frameingshell, text="Heat transfer coeff.", font=("Helvetica", 10))
    lb2.grid(row=8, column=1)
    lb2 = tk.Label(frameingshell, text=str(h2), font=("Helvetica", 10))
    lb2.grid(row=8, column=2)
    lb2 = tk.Label(frameingshell, text="W/m2C", font=("Helvetica", 10))
    lb2.grid(row=8, column=3)
    lb2 = tk.Label(frameingshell, text="Pressure", font=("Helvetica", 10))
    lb2.grid(row=9, column=1)
    lb2 = tk.Label(frameingshell, text=str(pressure2), font=("Helvetica", 10))
    lb2.grid(row=9, column=2)
    lb2 = tk.Label(frameingshell, text="kN/m2", font=("Helvetica", 10))
    lb2.grid(row=9, column=3)
    lb2 = tk.Label(frameingshell, text="Bundle Diameter", font=("Helvetica", 10))
    lb2.grid(row=2, column=1)
    lb2 = tk.Label(frameingshell, text=str(bundledia), font=("Helvetica", 10))
    lb2.grid(row=2, column=2)
    lb2 = tk.Label(frameingshell, text="Nusset No.", font=("Helvetica", 10))
    lb2.grid(row=11, column=1)
    lb2 = tk.Label(frameingshell, text=str(Nu2), font=("Helvetica", 10))
    lb2.grid(row=11, column=2)

    # -------------------------------Final Overall Calculation
    lb2 = tk.Label(frameingoverall, text="U(calculated)", font=("Helvetica", 10))
    lb2.grid(row=11, column=1)
    lb2 = tk.Label(frameingoverall, text=str(Ucal), font=("Helvetica", 10))
    lb2.grid(row=11, column=2)
    lb2 = tk.Label(frameingoverall, text="W/m2.⁰C", font=("Helvetica", 10))
    lb2.grid(row=11, column=3)
    lb2 = tk.Label(frameingoverall, text="Error", font=("Helvetica", 10))
    lb2.grid(row=12, column=1)
    lb2 = tk.Label(frameingoverall, text=str(error), font=("Helvetica", 10))
    lb2.grid(row=12, column=2)


# ------------------------ Calculating ldirratio
def ldirratio():
    u = float(ind1.get())
    v = float(length.get())
    w = v / u
    Lid.delete(0, len(str(float(w))))
    Lid.insert(0, str(w))


# ------------------------------------- Velocity Correction
def velcorrect():
    tabcorr = tk.Toplevel()
    tabcorr.title("correction")
    tabcorr.geometry("800x550")

    def calcrey():
        def correcting():
            text4 = tk.Label(tabcorr, text="velocity after", font=("Helvetica", 10))
            text4.grid(row=10, column=1)
            vtube2 = tk.Entry(tabcorr, width=20, borderwidth=5)
            vtube2.grid(row=10, column=2)
            text4 = tk.Label(tabcorr, text="m/s", font=("Helvetica", 10))
            text4.grid(row=10, column=3)
            text4 = tk.Label(tabcorr, text="velocity after", font=("Helvetica", 10))
            text4.grid(row=10, column=4)
            vshell2 = tk.Entry(tabcorr, width=20, borderwidth=5)
            vshell2.grid(row=10, column=5)
            text4 = tk.Label(tabcorr, text="m/s", font=("Helvetica", 10))
            text4.grid(row=10, column=6)
            text4 = tk.Label(tabcorr, text="coeff of heat after", font=("Helvetica", 10))
            text4.grid(row=11, column=1)
            heat1 = tk.Entry(tabcorr, width=20, borderwidth=5)
            heat1.grid(row=11, column=2)
            text4 = tk.Label(tabcorr, text="W/m2C", font=("Helvetica", 10))
            text4.grid(row=11, column=3)
            text4 = tk.Label(tabcorr, text="coeff of heat after", font=("Helvetica", 10))
            text4.grid(row=11, column=4)
            heat2 = tk.Entry(tabcorr, width=20, borderwidth=5)
            heat2.grid(row=11, column=5)
            text4 = tk.Label(tabcorr, text="W/m2C", font=("Helvetica", 10))
            text4.grid(row=11, column=6)
            text4 = tk.Label(tabcorr, text="pressure after", font=("Helvetica", 10))
            text4.grid(row=12, column=1)
            press1 = tk.Entry(tabcorr, width=20, borderwidth=5)
            press1.grid(row=12, column=2)
            text4 = tk.Label(tabcorr, text="kN/m2", font=("Helvetica", 10))
            text4.grid(row=12, column=3)
            text4 = tk.Label(tabcorr, text="pressure after", font=("Helvetica", 10))
            text4.grid(row=12, column=4)
            press2 = tk.Entry(tabcorr, width=20, borderwidth=5)
            press2.grid(row=12, column=5)
            text4 = tk.Label(tabcorr, text="kN/m2", font=("Helvetica", 10))
            text4.grid(row=12, column=6)
            text4 = tk.Label(tabcorr, text="U(new)", font=("Helvetica", 10))
            text4.grid(row=13, column=4)
            unew = tk.Entry(tabcorr, width=20, borderwidth=5)
            unew.grid(row=13, column=5)
            text4 = tk.Label(tabcorr, text="W", font=("Helvetica", 10))
            text4.grid(row=13, column=6)
            text4 = tk.Label(tabcorr, text="Error", font=("Helvetica", 10))
            text4.grid(row=13, column=1)
            error1 = tk.Entry(tabcorr, width=20, borderwidth=5)
            error1.grid(row=13, column=2)

            # tube
            mflow1 = (float(mflow2.get()) * float(spheat2.get()) * (float(inshell.get()) - float(outshell.get()))) / (
                    (float(outtube.get()) - float(intube.get())) * float(spheat1.get()))
            heatflow1 = mflow1 * float(spheat1.get()) * (float(outtube.get()) - float(intube.get()))
            delt = ((float(inshell.get()) - float(outtube.get())) - (float(outshell.get()) - float(intube.get()))) / (
                np.log((float(inshell.get()) - float(outtube.get())) / (float(outshell.get()) - float(intube.get()))))
            areatube = heatflow1 / (float(uass.get()) * float(fowling.get()) * delt * 3600)
            area1 = (22 * (float(ind1.get()) ** 2)) / 28
            nooftube = (areatube * 7) / (22 * float(outd1.get()) * float(length.get()))
            velocity1 = (mflow1 * float(nofp.get())) / (nooftube * float(den1.get()) * area1 * 3600)
            reynold1 = (float(ind1.get()) * velocity1 * float(den1.get())) / float(vis1.get()) * float(vtube1.get())
            pr1 = (float(vis1.get()) * float(spheat1.get())) / float(tcon1.get())

            h1 = (float(tcon1.get()) * reynold1 * (pr1 ** 0.33) * float(jhtnew.get())) / float(ind1.get())
            pressure1 = float(nofp.get()) * (
                    ((8 * float(jftnew.get()) * (float(length.get()) / float(ind1.get()))) + 2.5) * (
                    float(den1.get()) * (velocity1 ** 2))) / 2000

            # shell
            heatflow2 = float(mflow2.get()) * float(spheat2.get()) * (float(inshell.get()) - float(outshell.get()))
            bundledia = float(outd1.get()) * ((nooftube / float(k1s.get())) ** (1 / float(n1s.get())))
            pitch = 1.25 * float(outd1.get())
            shelldia = 1.1 * bundledia
            bafflespace = shelldia / 5
            area2 = ((pitch - float(outd1.get())) * shelldia * bafflespace) / pitch
            eqdia = 0
            if clickpitch.get() == "Triangular pitch":
                eqdia = 1.1 * ((pitch ** 2) - (0.917 * (float(outd1.get()) ** 2))) / float(outd1.get())
            elif clickpitch.get() == "Square pitch":
                eqdia = 1.27 * ((pitch ** 2) - (0.785 * (float(outd1.get()) ** 2))) / float(outd1.get())
            velocity2 = float(mflow2.get()) / (float(den2.get()) * area2 * 3600)
            reynold2 = (eqdia * velocity2 * float(den2.get())) / (float(vis2.get())) * float(vshell1.get())
            pr2 = float(vis2.get()) * float(spheat2.get()) / (float(tcon2.get()))
            h2 = (float(tcon2.get()) * reynold2 * (pr2 ** 0.33) * float(jhsnew.get())) / eqdia
            pressure2 = (8 * float(jfsnew.get()) * shelldia * float(length.get()) * float(den2.get()) * (
                    velocity2 ** 2)) / (eqdia * bafflespace * 2000)

            Ucal1 = (((1 / h1) * (float(outd1.get()) / float(ind1.get()))) + (1 / h2) + (
                    (float(outd1.get()) * np.log(float(outd1.get()) / float(ind1.get()))) / (
                    2 * float(alloy.get()))) + float(fouls.get()) + (
                             float(foult.get()) * (float(outd1.get()) / float(ind1.get())))) ** (-1)
            error = ((Ucal1 - float(uass.get())) / Ucal1) * 100

            newvel1 = velocity1 * float(vtube1.get())
            newvel2 = velocity2 * float(vshell1.get())
            h1n = h1 * ((float(vtube1.get())) ** 0.8)
            h2n = h2 * ((float(vshell1.get())) ** 0.8)
            pre1 = ((newvel1 / velocity1) ** 2) * pressure1
            pre2 = ((newvel2 / velocity2) ** 2) * pressure2
            unew1 = (((1 / h1n) * (float(outd1.get()) / float(ind1.get()))) + (1 / h2n) + (
                    (float(outd1.get()) * np.log(float(outd1.get()) / float(ind1.get()))) / (
                    2 * float(alloy.get()))) + float(fouls.get()) + (
                             float(foult.get()) * (float(outd1.get()) / float(ind1.get())))) ** (-1)

            error1.delete(0, len(str(float(error))))
            error1.insert(0, str(error))
            vtube2.delete(0, len(str(float(newvel1))))
            vtube2.insert(0, str(newvel1))
            vshell2.delete(0, len(str(float(newvel2))))
            vshell2.insert(0, str(newvel2))
            heat1.delete(0, len(str(float(h1n))))
            heat1.insert(0, str(h1n))
            heat2.delete(0, len(str(float(h2n))))
            heat2.insert(0, str(h2n))
            press1.delete(0, len(str(float(pre1))))
            press1.insert(0, str(pre1))
            press2.delete(0, len(str(float(pre2))))
            press2.insert(0, str(pre2))
            unew.delete(0, len(str(float(unew1))))
            unew.insert(0, str(unew1))

        text4 = tk.Label(tabcorr, text="Re(new)", font=("Helvetica", 10))
        text4.grid(row=6, column=1)
        rnewt = tk.Entry(tabcorr, width=20, borderwidth=5)
        rnewt.grid(row=6, column=2)
        text4 = tk.Label(tabcorr, text="Re(new)", font=("Helvetica", 10))
        text4.grid(row=6, column=4)
        rnews = tk.Entry(tabcorr, width=20, borderwidth=5)
        rnews.grid(row=6, column=5)

        text4 = tk.Label(tabcorr, text="Jh(new)", font=("Helvetica", 10))
        text4.grid(row=7, column=1)
        jhtnew = tk.Entry(tabcorr, width=20, borderwidth=5)
        jhtnew.grid(row=7, column=2)
        text4 = tk.Label(tabcorr, text="Jh(new)", font=("Helvetica", 10))
        text4.grid(row=7, column=4)
        jhsnew = tk.Entry(tabcorr, width=20, borderwidth=5)
        jhsnew.grid(row=7, column=5)

        text4 = tk.Label(tabcorr, text="Jf(new)", font=("Helvetica", 10))
        text4.grid(row=8, column=1)
        jftnew = tk.Entry(tabcorr, width=20, borderwidth=5)
        jftnew.grid(row=8, column=2)
        text4 = tk.Label(tabcorr, text="Jf(new)", font=("Helvetica", 10))
        text4.grid(row=8, column=4)
        jfsnew = tk.Entry(tabcorr, width=20, borderwidth=5)
        jfsnew.grid(row=8, column=5)

        # tube
        mflow1 = (float(mflow2.get()) * float(spheat2.get()) * (float(inshell.get()) - float(outshell.get()))) / (
                (float(outtube.get()) - float(intube.get())) * float(spheat1.get()))
        heatflow1 = mflow1 * float(spheat1.get()) * (float(outtube.get()) - float(intube.get()))
        delt = ((float(inshell.get()) - float(outtube.get())) - (float(outshell.get()) - float(intube.get()))) / (
            np.log((float(inshell.get()) - float(outtube.get())) / (float(outshell.get()) - float(intube.get()))))
        areatube = heatflow1 / (float(uass.get()) * float(fowling.get()) * delt * 3600)
        area1 = (22 * (float(ind1.get()) ** 2)) / 28
        nooftube = (areatube * 7) / (22 * float(outd1.get()) * float(length.get()))
        velocity1 = (mflow1 * float(nofp.get())) / (nooftube * float(den1.get()) * area1 * 3600)
        reynold3 = (float(ind1.get()) * velocity1 * float(den1.get())) / float(vis1.get()) * float(vtube1.get())

        # shell
        heatflow2 = float(mflow2.get()) * float(spheat2.get()) * (float(inshell.get()) - float(outshell.get()))
        bundledia = float(outd1.get()) * ((nooftube / float(k1s.get())) ** (1 / float(n1s.get())))
        pitch = 1.25 * float(outd1.get())
        shelldia = 1.1 * bundledia
        bafflespace = shelldia / 5
        area2 = ((pitch - float(outd1.get())) * shelldia * bafflespace) / pitch
        eqdia = 0
        if clickpitch.get() == "Triangular pitch":
            eqdia = 1.1 * ((pitch ** 2) - (0.917 * (float(outd1.get()) ** 2))) / float(outd1.get())
        elif clickpitch.get() == "Square pitch":
            eqdia = 1.27 * ((pitch ** 2) - (0.785 * (float(outd1.get()) ** 2))) / float(outd1.get())
        velocity2 = float(mflow2.get()) / (float(den2.get()) * area2 * 3600)
        reynold4 = (eqdia * velocity2 * float(den2.get())) / (float(vis2.get())) * float(vshell1.get())

        rnewt.delete(0, len(str(float(reynold3))))
        rnewt.insert(0, str(reynold3))
        rnews.delete(0, len(str(float(reynold4))))
        rnews.insert(0, str(reynold4))
        f2btn = tk.Button(tabcorr, text="Calculate", padx=30, pady=10, command=correcting)
        f2btn.grid(row=9, column=3)

    # -----------------------------------Correction Text
    text4 = tk.Label(tabcorr, text="Correction", font=("Helvetica", 20))
    text4.grid(row=1, column=3)
    text4 = tk.Label(tabcorr, text="Tube side", font=("Helvetica", 20))
    text4.grid(row=2, column=2)
    text4 = tk.Label(tabcorr, text="Shell side", font=("Helvetica", 20))
    text4.grid(row=2, column=5)

    text4 = tk.Label(tabcorr, text="velocity change factor", font=("Helvetica", 10))
    text4.grid(row=4, column=1)
    vtube1 = tk.Entry(tabcorr, width=20, borderwidth=5)
    vtube1.grid(row=4, column=2)
    text4 = tk.Label(tabcorr, text="velocity change factor", font=("Helvetica", 10))
    text4.grid(row=4, column=4)
    vshell1 = tk.Entry(tabcorr, width=20, borderwidth=5)
    vshell1.grid(row=4, column=5)
    f2btn = tk.Button(tabcorr, text="Calculate reynold", padx=30, pady=10, command=calcrey)
    f2btn.grid(row=5, column=3)


# ------------------------Renolyds Calculation
def reyno():
    mflow1 = (float(mflow2.get()) * float(spheat2.get()) * (float(inshell.get()) - float(outshell.get()))) / (
            (float(outtube.get()) - float(intube.get())) * float(spheat1.get()))
    heatflow1 = mflow1 * float(spheat1.get()) * (float(outtube.get()) - float(intube.get()))
    delt = ((float(inshell.get()) - float(outtube.get())) - (float(outshell.get()) - float(intube.get()))) / (
        np.log((float(inshell.get()) - float(outtube.get())) / (float(outshell.get()) - float(intube.get()))))
    areatube = heatflow1 / (float(uass.get()) * float(fowling.get()) * delt * 3600)
    area1 = (22 * (float(ind1.get()) ** 2)) / 28
    nooftube = (areatube * 7) / (22 * float(outd1.get()) * float(length.get()))
    velocity1 = (mflow1 * float(nofp.get())) / (nooftube * float(den1.get()) * area1 * 3600)
    reynold1 = (float(ind1.get()) * velocity1 * float(den1.get())) / float(vis1.get())

    heatflow2 = float(mflow2.get()) * float(spheat2.get()) * (float(inshell.get()) - float(outshell.get()))
    bundledia = float(outd1.get()) * ((nooftube / float(k1s.get())) ** (1 / float(n1s.get())))
    pitch = 1.25 * float(outd1.get())
    shelldia = 1.1 * bundledia
    bafflespace = shelldia / 5
    area2 = ((pitch - float(outd1.get())) * shelldia * bafflespace) / pitch
    eqdia = 0
    if clickpitch.get() == "Triangular pitch":
        eqdia = 1.1 * ((pitch ** 2) - (0.917 * (float(outd1.get()) ** 2))) / float(outd1.get())
    elif clickpitch.get() == "Square pitch":
        eqdia = 1.27 * ((pitch ** 2) - (0.785 * (float(outd1.get()) ** 2))) / float(outd1.get())
    velocity2 = float(mflow2.get()) / (float(den2.get()) * area2 * 3600)
    reynold2 = (eqdia * velocity2 * float(den2.get())) / (float(vis2.get()))

    ret.delete(0, len(str(float(reynold1))))
    ret.insert(0, str(reynold1))
    res.delete(0, len(str(float(reynold2))))
    res.insert(0, str(reynold2))


# ---------------------------------------R S Calculation
def crop():
    R = (float(inshell.get()) - float(outshell.get())) / (float(outtube.get()) - float(intube.get()))
    S = (float(outtube.get()) - float(intube.get())) / (float(inshell.get()) - float(outtube.get()))
    rvalue.delete(0, len(str(float(R))))
    rvalue.insert(0, str(R))
    svalue.delete(0, len(str(float(S))))
    svalue.insert(0, str(S))


# -----------------------------Showing Image in Ft
def showpic():
    if passing.get() == "2":
        img = Image.open("2pass.png")
        img = ImageTk.PhotoImage(img)
        lbl.configure(image=img)
        lbl.image = img
    elif passing.get() == "4":
        img = Image.open("4pass.png")
        img = ImageTk.PhotoImage(img)
        lbl.configure(image=img)
        lbl.image = img


lbl = tk.Label(frame11)
lbl.pack()

f2btn = tk.Button(frame11, text="Back", padx=30, pady=10, command=lambda: show_frame(frame7))
f2btn.place(x=40, y=400)

f2btn = tk.Button(frame11, text="Back", padx=30, pady=10, command=lambda: show_frame(frame6))
f2btn.place(x=1000, y=400)

f2btn = tk.Button(frame7, text="Back", padx=30, pady=10, command=lambda: show_frame(frame2))
f2btn.grid(row=20, column=1)

y = tk.Label(frame7, text="Tube Side", font=("Helvetica", 10))
y.grid(row=1, column=2)
y = tk.Label(frame7, text="Shell Side", font=("Helvetica", 10))
y.grid(row=1, column=6)

# ------------------------------Tube Side Parameters Present in Frameing 1
text4 = tk.Label(frame7, text="Enter inlet temp.", font=("Helvetica", 10))
text4.grid(row=2, column=1)
intube = tk.Entry(frame7, width=20, borderwidth=5)
intube.grid(row=2, column=3)
text4 = tk.Label(frame7, text="C", font=("Helvetica", 10))
text4.grid(row=2, column=4)
text4 = tk.Label(frame7, text="C", font=("Helvetica", 10))
text4.grid(row=3, column=4)
text4 = tk.Label(frame7, text="Enter outlet temp.", font=("Helvetica", 10))
text4.grid(row=3, column=1)
outtube = tk.Entry(frame7, width=20, borderwidth=5)
outtube.grid(row=3, column=3)
text4 = tk.Label(frame7, text="Specific heat", font=("Helvetica", 10))
text4.grid(row=4, column=1)
spheat1 = tk.Entry(frame7, width=20, borderwidth=5)
spheat1.grid(row=4, column=3)
text4 = tk.Label(frame7, text="j/Kgk", font=("Helvetica", 10))
text4.grid(row=4, column=4)
text4 = tk.Label(frame7, text="Density", font=("Helvetica", 10))
text4.grid(row=5, column=1)
den1 = tk.Entry(frame7, width=20, borderwidth=5)
den1.grid(row=5, column=3)
text4 = tk.Label(frame7, text="Kg/m3", font=("Helvetica", 10))
text4.grid(row=5, column=4)
text4 = tk.Label(frame7, text="Dynamic Viscosity", font=("Helvetica", 10))
text4.grid(row=6, column=1)
vis1 = tk.Entry(frame7, width=20, borderwidth=5)
vis1.grid(row=6, column=3)
text4 = tk.Label(frame7, text="Ns/m2", font=("Helvetica", 10))
text4.grid(row=6, column=4)
text4 = tk.Label(frame7, text="Thermal Conductivity", font=("Helvetica", 10))
text4.grid(row=7, column=1)
tcon1 = tk.Entry(frame7, width=20, borderwidth=5)
tcon1.grid(row=7, column=3)
text4 = tk.Label(frame7, text="W/m.⁰C", font=("Helvetica", 10))
text4.grid(row=7, column=4)
text4 = tk.Label(frame7, text="Jh(Tube)", font=("Helvetica", 10))
text4.grid(row=8, column=1)
jht = tk.Entry(frame7, width=20, borderwidth=5)
jht.grid(row=8, column=3)
text4 = tk.Label(frame7, text="Jf(Tube)", font=("Helvetica", 10))
text4.grid(row=9, column=1)
jft = tk.Entry(frame7, width=20, borderwidth=5)
jft.grid(row=9, column=3)
text4 = tk.Label(frame7, text="fouling(Tube)", font=("Helvetica", 10))
text4.grid(row=10, column=1)
foult = tk.Entry(frame7, width=20, borderwidth=5)
foult.grid(row=10, column=3)
text4 = tk.Label(frame7, text="Thermal Conduct of Alloy", font=("Helvetica", 10))
text4.grid(row=11, column=1)
alloy = tk.Entry(frame7, width=20, borderwidth=5)
alloy.grid(row=11, column=3)
text4 = tk.Label(frame7, text="W/m C", font=("Helvetica", 10))
text4.grid(row=11, column=4)

# ---------------------------------------Shell side Parameters Frameing 1

text4 = tk.Label(frame7, text="Enter inlet temp.", font=("Helvetica", 10))
text4.grid(row=2, column=5)
inshell = tk.Entry(frame7, width=20, borderwidth=5)
inshell.grid(row=2, column=7)
text4 = tk.Label(frame7, text="Enter outlet temp.", font=("Helvetica", 10))
text4.grid(row=3, column=5)
outshell = tk.Entry(frame7, width=20, borderwidth=5)
outshell.grid(row=3, column=7)
text4 = tk.Label(frame7, text="C", font=("Helvetica", 10))
text4.grid(row=2, column=8)
text4 = tk.Label(frame7, text="C", font=("Helvetica", 10))
text4.grid(row=3, column=8)
text4 = tk.Label(frame7, text="Specific heat", font=("Helvetica", 10))
text4.grid(row=4, column=5)
spheat2 = tk.Entry(frame7, width=20, borderwidth=5)
spheat2.grid(row=4, column=7)
text4 = tk.Label(frame7, text="j/Kgk", font=("Helvetica", 10))
text4.grid(row=4, column=8)
text4 = tk.Label(frame7, text="Density", font=("Helvetica", 10))
text4.grid(row=5, column=5)
den2 = tk.Entry(frame7, width=20, borderwidth=5)
den2.grid(row=5, column=7)
text4 = tk.Label(frame7, text="Kg/m3", font=("Helvetica", 10))
text4.grid(row=5, column=8)
text4 = tk.Label(frame7, text="Dynamic Viscosity", font=("Helvetica", 10))
text4.grid(row=6, column=5)
vis2 = tk.Entry(frame7, width=20, borderwidth=5)
vis2.grid(row=6, column=7)
text4 = tk.Label(frame7, text="Ns/m2", font=("Helvetica", 10))
text4.grid(row=6, column=8)
text4 = tk.Label(frame7, text="Thermal Conductivity", font=("Helvetica", 10))
text4.grid(row=7, column=5)
tcon2 = tk.Entry(frame7, width=20, borderwidth=5)
tcon2.grid(row=7, column=7)
text4 = tk.Label(frame7, text="W/m.⁰C", font=("Helvetica", 10))
text4.grid(row=7, column=8)
text4 = tk.Label(frame7, text="kg/hr", font=("Helvetica", 10))
text4.grid(row=8, column=8)
text4 = tk.Label(frame7, text="Mass flow rate", font=("Helvetica", 10))
text4.grid(row=8, column=5)
mflow2 = tk.Entry(frame7, width=20, borderwidth=5)
mflow2.grid(row=8, column=7)
text4 = tk.Label(frame7, text="Jh(shell)", font=("Helvetica", 10))
text4.grid(row=9, column=5)
jhs = tk.Entry(frame7, width=20, borderwidth=5)
jhs.grid(row=9, column=7)
text4 = tk.Label(frame7, text="Jf(shell)", font=("Helvetica", 10))
text4.grid(row=10, column=5)
jfs = tk.Entry(frame7, width=20, borderwidth=5)
jfs.grid(row=10, column=7)
text4 = tk.Label(frame7, text="fouling(shell)", font=("Helvetica", 10))
text4.grid(row=11, column=5)
fouls = tk.Entry(frame7, width=20, borderwidth=5)
fouls.grid(row=11, column=7)

# -----Common Part Parameters on Frameing 1

text4 = tk.Label(frame7, text="m", font=("Helvetica", 10))
text4.grid(row=2, column=11)
text4 = tk.Label(frame7, text="Outer tube diameter", font=("Helvetica", 10))
text4.grid(row=2, column=9)
outd1 = tk.Entry(frame7, width=20, borderwidth=5)
outd1.grid(row=2, column=10)
text4 = tk.Label(frame7, text="m", font=("Helvetica", 10))
text4.grid(row=3, column=11)
text4 = tk.Label(frame7, text="Inner tube diameter", font=("Helvetica", 10))
text4.grid(row=3, column=9)
ind1 = tk.Entry(frame7, width=20, borderwidth=5)
ind1.grid(row=3, column=10)
text4 = tk.Label(frame7, text="W/m2 C", font=("Helvetica", 10))
text4.grid(row=4, column=11)
text4 = tk.Label(frame7, text="U(assume)", font=("Helvetica", 10))
text4.grid(row=4, column=9)
uass = tk.Entry(frame7, width=20, borderwidth=5)
uass.grid(row=4, column=10)
text4 = tk.Label(frame7, text="No. of passes", font=("Helvetica", 10))
text4.grid(row=5, column=9)
nofp = tk.Entry(frame7, width=20, borderwidth=5)
nofp.grid(row=5, column=10)
text4 = tk.Label(frame7, text="Ft ", font=("Helvetica", 10))
text4.grid(row=6, column=9)
fowling = tk.Entry(frame7, width=20, borderwidth=5)
fowling.grid(row=6, column=10)
text4 = tk.Label(frame7, text="Length", font=("Helvetica", 10))
text4.grid(row=7, column=9)
length = tk.Entry(frame7, width=20, borderwidth=5)
length.grid(row=7, column=10)
text4 = tk.Label(frame7, text="m", font=("Helvetica", 10))
text4.grid(row=7, column=11)
text4 = tk.Label(frame7, text="K1(shell)", font=("Helvetica", 10))
text4.grid(row=9, column=9)
k1s = tk.Entry(frame7, width=20, borderwidth=5)
k1s.grid(row=9, column=10)
text4 = tk.Label(frame7, text="n1(shell)", font=("Helvetica", 10))
text4.grid(row=10, column=9)
n1s = tk.Entry(frame7, width=20, borderwidth=5)
n1s.grid(row=10, column=10)
f2btn = tk.Button(frame7, text="Coefficients", padx=30, pady=10, command=heatfac)
f2btn.grid(row=12, column=10)
text4 = tk.Label(frame7, text="Arrangement", font=("Helvetica", 10))
text4.grid(row=11, column=9)
approach = tk.OptionMenu(frame7, clickpitch, *clicktrsq)
approach.grid(row=11, column=10)

# --------------------------- l/di
text4 = tk.Label(frame7, text="L/di", font=("Helvetica", 10))
text4.grid(row=8, column=9)
Lid = tk.Entry(frame7, width=20, borderwidth=5)
Lid.grid(row=8, column=10)
f2btn = tk.Button(frame7, text="L/di ratio", padx=37, pady=10, command=ldirratio)
f2btn.grid(row=13, column=10)
# -------------------------Correction
f2btn = tk.Button(frame7, text="Correction", padx=30, pady=10, command=velcorrect)
f2btn.grid(row=14, column=5)

#  ---------------------------------Reynolds value on Frameing 1
text4 = tk.Label(frame7, text="reynold(shell)", font=("Helvetica", 10))
text4.grid(row=12, column=5)
res = tk.Entry(frame7, width=20, borderwidth=5)
res.grid(row=12, column=7)
text4 = tk.Label(frame7, text="reynold(tube)", font=("Helvetica", 10))
text4.grid(row=12, column=1)
ret = tk.Entry(frame7, width=20, borderwidth=5)
ret.grid(row=12, column=3)
f2btn = tk.Button(frame7, text="Value", padx=10, pady=5, command=reyno)
f2btn.grid(row=12, column=8)

# ------------------------------R & S Value
text4 = tk.Label(frame7, text="R", font=("Helvetica", 10))
text4.grid(row=13, column=5)
rvalue = tk.Entry(frame7, width=20, borderwidth=5)
rvalue.grid(row=13, column=7)
text4 = tk.Label(frame7, text="S", font=("Helvetica", 10))
text4.grid(row=13, column=1)
svalue = tk.Entry(frame7, width=20, borderwidth=5)
svalue.grid(row=13, column=3)
f2btn = tk.Button(frame7, text="Value", padx=10, pady=5, command=crop)
f2btn.grid(row=13, column=8)
# -----------------
f2btn = tk.Button(frame7, text="Ft value", padx=40, pady=10, command=lambda: show_frame(frame11))
f2btn.grid(row=14, column=10)

# -------------------------------Ft value Graph
passing = tk.StringVar()
choice = ["2", "4"]
passing.set(choice[0])
y = tk.Label(frame11, text="Choose no. of passes", font=("Helvetica", 10))
y.place(x=40, y=20)
choice1 = tk.OptionMenu(frame11, passing, *choice)
choice1.place(x=180, y=20)
f2btn = tk.Button(frame11, text="select", padx=20, pady=5, command=showpic)
f2btn.place(x=60, y=60)

# --------------- Final Calculation Button calling shellatube() Function
f2btn = tk.Button(frame7, text="Calculate", padx=30, pady=10, command=shellatube)
f2btn.grid(row=14, column=4)




# -------------------------------------------------------------------------Excel Optimization--------------------------------------------------------------------------


y = tk.Label(frame6, text="Tube Side", font=("Helvetica", 10))
y.grid(row=1, column=2)
y = tk.Label(frame6, text="Shell Side", font=("Helvetica", 10))
y.grid(row=1, column=6)


# ------------------------------ R and S for Excel
def crop1():
    R = (float(inshellexcel.get()) - float(outshellexcel.get())) / (
                float(outtubeexcel.get()) - float(intubeexcel.get()))
    S = (float(outtubeexcel.get()) - float(intubeexcel.get())) / (float(inshellexcel.get()) - float(outtubeexcel.get()))
    rvalueexcel.delete(0, len(str(float(R))))
    rvalueexcel.insert(0, str(R))
    svalueexcel.delete(0, len(str(float(S))))
    svalueexcel.insert(0, str(S))


# -------------------------------- Calculating Reynold No. For Excel

# Pitch Triangular and Square from excel
columnP = ws['C8':'C47']
listsP = []
for cell in columnP:
    for x in cell:
        intemp = x.value
        listsP.append(intemp)


# Outer Diameter from excel
columnOd = ws['E8':'E47']
listsdia = []
for cell in columnOd:
    for x in cell:
        intemp = x.value
        listsdia.append(intemp)


# Thickness from excel
columnTh = ws['F8':'F47']
listsTh = []
for cell in columnTh:
    for x in cell:
        intemp = x.value
        listsTh.append(intemp)


# Length from excel
columnL = ws['H8':'H47']
listsL = []
for cell in columnL:
    for x in cell:
        intemp = x.value
        listsL.append(intemp)



# --------------- Tube Side

# No. Of Passes from excel
columnNp = ws['L8':'L47']
listsNp = []
for cell in columnNp:
    for x in cell:
        intemp = x.value
        listsNp.append(intemp)


# ----------------------Shell Side
# Baffle Cut from excel
columnBc = ws['V8':'V47']
listsBc = []
for cell in columnBc:
    for x in cell:
        intemp = x.value
        listsBc.append(intemp)


# --------------Overall Values
# Overall HT Uass from excel
columnU = ws['AL8':'AL47']
listsU = []
for cell in columnU:
    for x in cell:
        intemp = x.value
        listsU.append(intemp)
# print(listsU)

# ------------------------------ Inner Dia  And L/ID  ratio

#-------------------- Inner Dia
listsIdia = []
for i in range(0,40):
    ID = listsdia[i] - (2 * listsTh[i])
    listsIdia.append(ID)
for k in range(0, 40):
    put = listsIdia[k]
    ws.cell(row=k + 8, column=7).value = put
# ------------------------- L/ID ratio
listsLid = []
for i in range(0,40):
    lid = listsL[i] / listsIdia[i]
    listsLid.append(lid)
for k in range(0, 40):
    put = listsLid[k]
    ws.cell(row=k + 8, column=9).value = put



def shellandtubeExcel():

# -------------------------- For Excel Tube Side

    mflow1excel = (float(mflow2excel.get()) * float(spheat2excel.get()) * (
                float(inshellexcel.get()) - float(outshellexcel.get()))) / (
                          (float(outtubeexcel.get()) - float(intubeexcel.get())) * float(spheat1excel.get()))
    # print(mflow1excel)
    heatflow1excel = mflow1excel * float(spheat1excel.get()) * (float(outtubeexcel.get()) - float(intubeexcel.get()))
    # print(heatflow1excel)
    deltexcel = ((float(inshellexcel.get()) - float(outtubeexcel.get())) - (
                float(outshellexcel.get()) - float(intubeexcel.get()))) / (
                    np.log((float(inshellexcel.get()) - float(outtubeexcel.get())) / (
                                float(outshellexcel.get()) - float(intubeexcel.get()))))
    # print(deltexcel)

    #-------------------------------- Area Of Tube
    listAreaT = []
    for i in range(0,40):
        areaoftubeexcel = (22 * (listsIdia[i] ** 2)) / 28
        listAreaT.append(areaoftubeexcel)
    for k in range(0, 40):
        put = listAreaT[k]
        ws.cell(row=k + 8, column=13).value = put

        # ------------------ Area for Calculating No. of Tube
    listArea = []
    for i in range(0, 40):
        areatubeexcel = heatflow1excel / (listsU[i] * (float(fowlingexcel.get()) * deltexcel * 3600))
        listArea.append(areatubeexcel)
    # print("AreaTube values\n", listArea)
    # for k in range(0, 40):
    #     put = listArea[k]
    #     ws.cell(row=k + 8, column=7).value = put
    # ------------------------No. of Tube
    listNT = []
    for i in range(0, 40):
        nooftubeexcel = (listArea[i] * 7) / (22 * listsdia[i] * listsL[i])
        listNT.append(nooftubeexcel)
    # print("No.Tubes values\n", listNT)
    for k in range(0, 40):
        put = listNT[k]
        ws.cell(row=k + 8, column=10).value = put

    # -----------------------------Tube Side Velocity
    listVel1 = []
    for i in range(0, 40):
        velocity1excel = (mflow1excel * 28 * listsNp[i]) / (listNT[i] * float(den1excel.get()) * 22 * (listsIdia[i] ** 2) * 3600)
        listVel1.append(velocity1excel)
    # print("Tube Velocity values\n", listVel1)
    for k in range(0, 40):
        put = listVel1[k]
        ws.cell(row=k + 8, column=14).value = put

    # --------------- Tube Side Reynolds No.
    listR1 = []
    for i in range(0, 40):
        reynold1excel = ((listsIdia[i]) * listVel1[i] * float(den1excel.get())) / float(vis1excel.get())
        listR1.append(reynold1excel)
    # print("Tube Side \n" , listR1)
    for k in range(0, 40):
        put = listR1[k]
        ws.cell(row=k + 8, column=15).value = put

# -------------------------- For Excel Shell Side


# -----------------------------Value for k1 and n1
    listsk1 = []
    listsn1 = []
    for i in range(0, 40):
        if (listsP[i] == "Triangular Pitch" and listsNp[i] == 1):
            u = 0.319
            v = 2.142
            listsk1.append(u)
            listsn1.append(v)

        elif (listsP[i] == "Triangular Pitch" and listsNp[i] == 2):
            u = 0.249
            v = 2.207
            listsk1.append(u)
            listsn1.append(v)
        elif (listsP[i] == "Triangular Pitch" and listsNp[i] == 4):
            u = 0.175
            v = 2.285
            listsk1.append(u)
            listsn1.append(v)
        elif (listsP[i] == "Triangular Pitch" and listsNp[i] == 6):
            u = 0.0743
            v = 2.499
            listsk1.append(u)
            listsn1.append(v)
        elif (listsP[i] == "Triangular Pitch" and listsNp[i] == 8):
            u = 0.0365
            v = 2.675
            listsk1.append(u)
            listsn1.append(v)

        elif (listsP[i] == "Square Pitch" and listsNp[i] == 1):
            u = 0.215
            v = 2.207
            listsk1.append(u)
            listsn1.append(v)
        elif (listsP[i] == "Square Pitch" and listsNp[i] == 2):
            u = 0.156
            v = 2.291
            listsk1.append(u)
            listsn1.append(v)
        elif (listsP[i] == "Square Pitch" and listsNp[i] == 4):
            u = 0.158
            v = 2.263
            listsk1.append(u)
            listsn1.append(v)
        elif (listsP[i] == "Square Pitch" and listsNp[i] == 6):
            u = 0.0402
            v = 2.617
            listsk1.append(u)
            listsn1.append(v)
        elif (listsP[i] == "Square Pitch" and listsNp[i] == 8):
            u = 0.0331
            v = 2.643
            listsk1.append(u)
            listsn1.append(v)

    # print("K1 values\n", listsk1)
    # print("N1 values\n", listsn1)
    for k in range(0, 40):
        put = listsk1[k]
        ws.cell(row=k + 8, column=25).value = put

    for k in range(0, 40):
        put = listsn1[k]
        ws.cell(row=k + 8, column=26).value = put


# -------------------------- For Excel Shell Side
# -------------Q of Shell side
    heatflow2excel = float(mflow2excel.get()) * float(spheat2excel.get()) * (
                float(inshellexcel.get()) - float(outshellexcel.get()))

    #-------------------------- Bundle Diameter
    listBdia = []
    for i in range(0, 40):
        bundlediaexcel = listsdia[i] * ((listNT[i] / listsk1[i]) ** (1 / listsn1[i]))
        listBdia.append(bundlediaexcel)
    # print("BndDia values\n", listBdia)
    for k in range(0, 40):
        put = listBdia[k]
        ws.cell(row=k + 8, column=27).value = put

    # --------------- Shell side Pitch
    listSP = []
    for i in range(0, 40):
        pitchexcel = 1.25 * (listsdia[i])
        listSP.append(pitchexcel)
    # print("SPitch values\n", listSP)
    for k in range(0, 40):
        put = listSP[k]
        ws.cell(row=k + 8, column=11).value = put
    # ---------- Shell Side Diameter
    listDs = []
    for i in range(0, 40):
        shelldiaexcel = 1.1 * listBdia[i]
        listDs.append(shelldiaexcel)
    # print("SDia values\n", listDs)
    for k in range(0, 40):
        put = listDs[k]
        ws.cell(row=k + 8, column=24).value = put
    # ----------------- Baffle Space
    listBs = []
    for i in range(0, 40):
        bafflespaceexcel = listDs[i] / 5
        listBs.append(bafflespaceexcel)
    # print("Baffle Spacing values\n", listBs)
    for k in range(0, 40):
        put = listBs[k]
        ws.cell(row=k + 8, column=23).value = put
    # ------------------- Shell Side Area
    listAreaS = []
    for i in range(0, 40):
        area2excel = ((listSP[i] - listsdia[i]) * listDs[i] * listBs[i]) / listSP[i]
        listAreaS.append(area2excel)
    # print("AreaS values\n", listAreaS)
    for k in range(0, 40):
        put = listAreaS[k]
        ws.cell(row=k + 8, column=29).value = put
    # ------------------------ Equivalent Diameter
    listEdia = []
    for i in range(0, 40):
        if (listsP[i] == "Triangular Pitch"):
            eqdiaexcel = 1.1 * ((listSP[i] ** 2) - (0.917 * (listsdia[i] ** 2))) / listsdia[i]
            listEdia.append(eqdiaexcel)
        elif (listsP[i] == "Square Pitch"):
            eqdiaexcel = 1.27 * ((listSP[i] ** 2) - (0.785 * (listsdia[i] ** 2)))/ listsdia[i]
            listEdia.append(eqdiaexcel)
    # print("EQDia values\n", listEdia)
    for k in range(0, 40):
        put = listEdia[k]
        ws.cell(row=k + 8, column=28).value = put
    # --------------------- Shell Side Velocity
    listVel2 = []
    for i in range(0, 40):
        velocity2excel = float(mflow2excel.get()) / (float(den2excel.get()) * listAreaS[i] * 3600)
        listVel2.append(velocity2excel)
    # print("Velocity Shell values\n", listVel2)
    for k in range(0, 40):
        put = listVel2[k]
        ws.cell(row=k + 8, column=30).value = put
    #----------------------------Shell Dide Renolds No.
    listR2 = []
    for i in range(0, 40):
        reynold2excel = (listEdia[i] * listVel2[i] * float(den2excel.get())) / (float(vis2excel.get()))
        listR2.append(reynold2excel)
    # print("Reyno Shell values\n", listR2)
    for k in range(0, 40):
        put = listR2[k]
        ws.cell(row=k + 8, column=31).value = put


# ------- Using Graph of Re and Jh For Diff L/Di
# ---------------------Excel Tubeside Jh
#     optionheat = ["24", "48", "120", "240", "500"]
    jhtexcel = []
    for i in range(0, 40):
        z = listR1[i]
        ld = listsLid[i]
        # yo = 0
        if (12 < ld < 36):
            if z <= 2000:
                yo = (-0.131 * (np.log(z))) + 1.00024
            elif z > 2000 and z <= 10000:
                yo = (-6.25 * (10 ** (-8)) * z) + 0.004625
            elif z > 10000 and z <= 1000000:
                yo = (-2.22 * (10 ** (-9)) * z) + 0.004022
            jhtexcel.append(yo)

        elif (36 < ld < 90):
            if z <= 2000:
                yo = (-4.83417 * (10 ** (-5)) * z) + 0.1346
            elif z > 2000 and z <= 10000:
                yo = (0.000000025 * z) + 0.00375
            elif z > 10000 and z <= 1000000:
                yo = (-2.22 * (10 ** (-9)) * z) + 0.004022
            jhtexcel.append(yo)

        elif (90 < ld < 180):
            if z <= 2000:
                yo = (-3.87437 * (10 ** (-5)) * z) + 0.0803874
            elif z > 2000 and z <= 10000:
                yo = (-7 * (10 ** (-11)) * z * z) + (9 * (10 ** (-7)) * z) + 0.0013
            elif z > 10000 and z <= 1000000:
                yo = (-2.22 * (10 ** (-9)) * z) + 0.004022
            jhtexcel.append(yo)

        elif (180 < ld < 370):
            if z <= 2000:
                yo = (-2.91457 * (10 ** (-5)) * z) + 0.06
            elif z > 2000 and z <= 10000:
                yo = (-1 * (10 ** (-10)) * z * z) + (2 * (10 ** (-6)) * z) - 0.0006
            elif z > 10000 and z <= 1000000:
                yo = (-2.22 * (10 ** (-9)) * z) + 0.004022
            jhtexcel.append(yo)

        elif (370 < ld):
            if z <= 2000:
                yo = (-2.13065 * (10 ** (-5)) * z) + 0.04421
            elif z > 2000 and z <= 10000:
                yo = (-9 * (10 ** (-11)) * z * z) + (1 * (10 ** (-6)) * z) - 0.0007
            elif z > 10000 and z <= 1000000:
                yo = (-2.22 * (10 ** (-9)) * z) + 0.004022
            jhtexcel.append(yo)

    # print("Jh Tube\n", jhtexcel)
    for k in range(0, 40):
        put = jhtexcel[k]
        ws.cell(row=k + 8, column=16).value = put
# -------------- Excel Tubeside Jf

    jftexcel = []
    for i in range(0, 40):
        # yo = 0
        z = listR1[i]
        if z <= 2000:
            yo = (-0.0004 * z) + 0.804
        elif z > 2000 and z <= 3200:
            yo = (-1 * (10 ** (-9)) * z * z) + (8 * (10 ** (-6)) * z) - 0.0075
        elif z > 3200 and z <= 1000000:
            yo = (-4.3138 * (10 ** (-9)) * z) + 0.0061138

        jftexcel.append(yo)

    # print("JfTube\n", jftexcel)
    for k in range(0, 40):
        put = jftexcel[k]
        ws.cell(row=k + 8, column=17).value = put
#----------------------------------Excel Shell Side Jh

    jhsexcel = []
    for i in range(0, 40):
        z = listR2[i]

        if (listsBc[i] == 0.15) :
            if z <= 140:
                yo = (2 * (10 ** (-5)) * (z ** 2)) + ((-0.0039) * z) + 0.2513
            elif z > 140 and z <= 10000:
                yo = (7 * (10 ** (-10)) * (z ** 2)) + (-1 * (10 ** (-5)) * z) + 0.0303
            elif z > 10000 and z <= 100000:
                yo = (1 * (10 ** (-12)) * (z ** 2)) + (-2 * (10 ** (-7)) * z) + 0.008
            elif z > 100000 and z <= 1000000:
                yo = (-1.4444 * (10 ** (-9)) * z) + 0.002704
            jhsexcel.append(yo)


        elif (listsBc[i] == 0.25):
            if z <= 140:
                yo = (1 * (10 ** (-5)) * (z ** 2)) + ((-0.0032) * z) + 0.2257
            elif z > 140 and z <= 10000:
                yo = (4 * (10 ** (-9)) * (z ** 2)) + (-4 * (10 ** (-5)) * z) + 0.0464
            elif z > 10000 and z <= 100000:
                yo = (7 * (10 ** (-13)) * (z ** 2)) + (-1 * (10 ** (-7)) * z) + 0.0069
            elif z > 100000 and z <= 1000000:
                yo = (-1.44444 * (10 ** (-9)) * z) + 0.0021444
            jhsexcel.append(yo)

        elif (listsBc[i] == 0.35):
            if z <= 140:
                yo = (1 * (10 ** (-5)) * (z ** 2)) + ((-0.0029) * z) + 0.1863
            elif z > 140 and z <= 10000:
                yo = (9 * (10 ** (-10)) * (z ** 2)) + (-1 * (10 ** (-5)) * z) + 0.0351
            elif z > 10000 and z <= 100000:
                yo = (9 * (10 ** (-13)) * (z ** 2)) + (-1 * (10 ** (-7)) * z) + 0.0063
            elif z > 100000 and z <= 1000000:
                yo = (-1.4444 * (10 ** (-9)) * z) + 0.0020444
            jhsexcel.append(yo)

        elif (listsBc[i] == 0.45):
            if z <= 140:
                yo = (1 * (10 ** (-5)) * (z ** 2)) + ((-0.0031) * z) + 0.1874
            elif z > 140 and z <= 10000:
                yo = (7 * (10 ** (-10)) * (z ** 2)) + (-1 * (10 ** (-5)) * z) + 0.032
            elif z > 10000 and z <= 100000:
                yo = (5 * (10 ** (-13)) * (z ** 2)) + (-8 * (10 ** (-8)) * z) + 0.0058
            elif z > 100000 and z <= 1000000:
                yo = (-1.4444 * (10 ** (-9)) * z) + 0.0019422
            jhsexcel.append(yo)

    # print("JhShell Excell values\n", jhsexcel)
    for k in range(0, 40):
        put = jhsexcel[k]
        ws.cell(row=k + 8, column=32).value = put

#---------------------------------------------------- Shell Side Jf

    jfsexcel = []
    for i in range(0, 40):
        z = listR2[i]
        # yo = 0
        if (listsBc[i] == 0.15):
                if z >= 10 and z <= 300:
                    yo = (0.0002 * (z ** 2)) + (-0.074 * z) + 4.3203
                elif z > 300 and z <= 4000:
                    yo = (1 * (10 ** (-8)) * (z ** 2)) + (-6 * (10 ** (-5)) * z) + 0.1609
                elif z > 4000 and z <= 100000:
                    yo = (1 * (10 ** (-11)) * (z ** 2)) + (-1 * (10 ** (-6)) * z) + 0.0844
                elif z > 100000 and z <= 1000000:
                    yo = (2 * (10 ** (-14)) * (z ** 2)) + (-4 * (10 ** (-8)) * z) + 0.0509
                jfsexcel.append(yo)

        elif (listsBc[i] == 0.25):
                if z >= 10 and z <= 300:
                    yo = (9 * (10 ** (-5)) * (z ** 2)) + (-0.0349 * z) + 2.2396
                elif z > 300 and z <= 4000:
                    yo = (1 * (10 ** (-8)) * (z ** 2)) + (-6 * (10 ** (-5)) * z) + 0.114
                elif z > 4000 and z <= 100000:
                    yo = (4 * (10 ** (-12)) * (z ** 2)) + (-7 * (10 ** (-7)) * z) + 0.0592
                elif z > 100000 and z <= 1000000:
                    yo = (3 * (10 ** (-14)) * (z ** 2)) + (-4 * (10 ** (-8)) * z) + 0.0391
                jfsexcel.append(yo)

        elif (listsBc[i] == 0.35):
                if z >= 10 and z <= 300:
                    yo = (0.0001 * (z ** 2)) + (-0.048 * z) + 2.4668
                elif z > 300 and z <= 4000:
                    yo = (6 * (10 ** (-8)) * (z ** 2)) + (-0.0001 * z) + 0.1216
                elif z > 4000 and z <= 100000:
                    yo = (3 * (10 ** (-12)) * (z ** 2)) + (-6 * (10 ** (-7)) * z) + 0.0508
                elif z > 100000 and z <= 1000000:
                    yo = (2 * (10 ** (-14)) * (z ** 2)) + (-3 * (10 ** (-8)) * z) + 0.0312
                jfsexcel.append(yo)

        elif (listsBc[i] == 0.45):
                if z >= 10 and z <= 300:
                    yo = (5 * (10 ** (-5)) * (z ** 2)) + (-0.0195 * z) + 1.6924
                elif z > 300 and z <= 4000:
                    yo = (7 * (10 ** (-9)) * (z ** 2)) + (-4 * (10 ** (-5)) * z) + 0.0787
                elif z > 4000 and z <= 100000:
                    yo = (2 * (10 ** (-12)) * (z ** 2)) + (-4 * (10 ** (-7)) * z) + 0.0398
                elif z > 100000 and z <= 1000000:
                    yo = (2 * (10 ** (-14)) * (z ** 2)) + (-3 * (10 ** (-8)) * z) + 0.0261
                jfsexcel.append(yo)

    # print("JfShell Excel",jfsexcel)
    for k in range(0, 40):
        put = jfsexcel[k]
        ws.cell(row=k + 8, column=33).value = put

        # ----------------------------- Final TubeSIde Calculation

    # --------------------- Tube Side Prandlt No.
    listPn = []
    for i in range(0,40):
        pr1excel = (float(vis1excel.get()) * float(spheat1excel.get())) / float(tcon1excel.get())
        listPn.append(pr1excel)
    # print("Prandlt no." , listPn)
    for k in range(0, 40):
        put = listPn[k]
        ws.cell(row=k + 8, column=18).value = put
    # ------------------------ Tube Side Nusselt No.
    listNu1 = []
    for i in range(0, 40):
        Nu1excel = listR1[i] * jhtexcel[i] * (listPn[i] ** 0.33)
        listNu1.append(Nu1excel)
    # print("Nu1 Tube Excel", listNu1)
    for k in range(0, 40):
        put = listNu1[k]
        ws.cell(row=k + 8, column=19).value = put

    #------------------------- Tube Side Heat Transfer Coefficient
    listh1 = []
    for i in range(0, 40):
        # h1excel = (float(tcon1excel.get()) * listR1[i] * (listPn[i] ** 0.33) * jhtexcel[i]) / listsIdia[i]
        h1excel = (listNu1[i] * float(tcon1excel.get())) / listsIdia[i]
        listh1.append(h1excel)
    # print("H Tube Excel", listh1)
    for k in range(0, 40):
        put = listh1[k]
        ws.cell(row=k + 8, column=20).value = put


    listpressure1 = []
    for i in range(0, 40):
        pressure1 = (listsNp[i]) * (((8 * (jftexcel[i]) * (listsL[i] / listsIdia[i])) + 2.5) * (float(den1excel.get()) * (listVel1[i] ** 2))) / 2000
        listpressure1.append(pressure1)
    # print("Pressure Tube Excel", listpressure1)
    for k in range(0, 40):
        put = listpressure1[k]
        ws.cell(row=k + 8, column=21).value = put



# ------------------------------Final Shell Side Calculation

    listSpn = []
    for i in range(0,40):
        pr2excel = float(vis2excel.get()) * float(spheat2excel.get()) / (float(tcon2excel.get()))
        listSpn.append(pr2excel)
    # print(listSpn)
    for k in range(0, 40):
        put = listSpn[k]
        ws.cell(row=k + 8, column=34).value = put


# ------------------------ Shell Side Nusselt No.
    listNu2 = []
    for i in range(0, 40):
        Nu2excel = listR2[i] * jhsexcel[i] * (listSpn[i] ** 0.33)
        # Nu2excel = (listh2[i] * (listsdia[i])) / (float(tcon2excel.get()))
        listNu2.append(Nu2excel)
    # print("Nu Shell Excel", listNu2)
    for k in range(0, 40):
        put = listNu2[k]
        ws.cell(row=k + 8, column=35).value = put

# ------------------------ Shell Side Heat Transfer
    listh2 = []
    for i in range(0, 40):
        h2excel = (listNu2[i] * float(tcon2excel.get()))  / listEdia[i]
        # h2excel = (float(tcon2excel.get()) * listR2[i] * (listSpn[i] ** 0.33) * (jhsexcel[i])) / listEdia[i]
        listh2.append(h2excel)
    # print("H Shell Excel", listh2)
    for k in range(0, 40):
        put = listh2[k]
        ws.cell(row=k + 8, column=36).value = put

    # -------------------------- Shell Side Pressure
    listpressure2 = []
    for i in range(0, 40):
        pressure2 = (8 * (jfsexcel[i]) * listDs[i] * listsL[i] * float(den2excel.get()) * (listVel2[i] ** 2)) / (listEdia[i] * listBs[i] * 2000)
        listpressure2.append(pressure2)
    for k in range(0, 40):
        put = listpressure2[k]
        ws.cell(row=k + 8, column=37).value = put

    # ------------------ OverAll Heat Tranfer Coefficient Calculation

    listUcal = []
    for i in range(0, 40):
        Ucalexcel = (((1 / listh1[i]) * (listsdia[i] / listsIdia[i])) + (1 / listh2[i]) + ((listsdia[i] * np.log(listsdia[i] / listsIdia[i])) / (2 * float(alloyexcel.get()))) + float(foulsexcel.get()) + (float(foultexcel.get()) * (listsdia[i] / (listsIdia[i])))) ** (-1)
        listUcal.append(Ucalexcel)
    for k in range(0, 40):
        put = listUcal[k]
        ws.cell(row=k + 8, column=39).value = put

    listerror = []
    for i in range(0, 40):
        errorexcel = ((listUcal[i] - listsU[i]) / listUcal[i]) * 100
        listerror.append(errorexcel)
    for k in range(0, 40):
        put = listerror[k]
        ws.cell(row=k + 8, column=40).value = put


    wb.save("C:/Users/rehan/PycharmProjects/pythontuts/Major Project/MPExcel2.xlsx")



# ------------------------Tube Side value on Frameing 2

text4 = tk.Label(frame6, text="Enter inlet temp.", font=("Helvetica", 10))
text4.grid(row=2, column=1)
intubeexcel = tk.Entry(frame6, width=20, borderwidth=5)
intubeexcel.grid(row=2, column=3)
text4 = tk.Label(frame6, text="C", font=("Helvetica", 10))
text4.grid(row=2, column=4)
text4 = tk.Label(frame6, text="C", font=("Helvetica", 10))
text4.grid(row=3, column=4)
text4 = tk.Label(frame6, text="Enter outlet temp.", font=("Helvetica", 10))
text4.grid(row=3, column=1)
outtubeexcel = tk.Entry(frame6, width=20, borderwidth=5)
outtubeexcel.grid(row=3, column=3)
text4 = tk.Label(frame6, text="Specific heat", font=("Helvetica", 10))
text4.grid(row=4, column=1)
spheat1excel = tk.Entry(frame6, width=20, borderwidth=5)
spheat1excel.grid(row=4, column=3)
text4 = tk.Label(frame6, text="j/Kgk", font=("Helvetica", 10))
text4.grid(row=4, column=4)
text4 = tk.Label(frame6, text="Density", font=("Helvetica", 10))
text4.grid(row=5, column=1)
den1excel = tk.Entry(frame6, width=20, borderwidth=5)
den1excel.grid(row=5, column=3)
text4 = tk.Label(frame6, text="Kg/m3", font=("Helvetica", 10))
text4.grid(row=5, column=4)
text4 = tk.Label(frame6, text="Dynamic Viscosity", font=("Helvetica", 10))
text4.grid(row=6, column=1)
vis1excel = tk.Entry(frame6, width=20, borderwidth=5)
vis1excel.grid(row=6, column=3)
text4 = tk.Label(frame6, text="Ns/m2", font=("Helvetica", 10))
text4.grid(row=6, column=4)
text4 = tk.Label(frame6, text="Thermal Conductivity", font=("Helvetica", 10))
text4.grid(row=7, column=1)
tcon1excel = tk.Entry(frame6, width=20, borderwidth=5)
tcon1excel.grid(row=7, column=3)
text4 = tk.Label(frame6, text="W/m.⁰C", font=("Helvetica", 10))
text4.grid(row=7, column=4)
text4 = tk.Label(frame6, text="fouling(Tube)", font=("Helvetica", 10))
text4.grid(row=8, column=1)
foultexcel = tk.Entry(frame6, width=20, borderwidth=5)
foultexcel.grid(row=8, column=3)
text4 = tk.Label(frame6, text="Thermal Conduct of Alloy", font=("Helvetica", 10))
text4.grid(row=9, column=1)
alloyexcel = tk.Entry(frame6, width=20, borderwidth=5)
alloyexcel.grid(row=9, column=3)
text4 = tk.Label(frame6, text="W/m C", font=("Helvetica", 10))
text4.grid(row=9, column=4)

# --------------------------Shell side values on Frameing 2

text4 = tk.Label(frame6, text="Enter inlet temp.", font=("Helvetica", 10))
text4.grid(row=2, column=5)
inshellexcel = tk.Entry(frame6, width=20, borderwidth=5)
inshellexcel.grid(row=2, column=7)
text4 = tk.Label(frame6, text="Enter outlet temp.", font=("Helvetica", 10))
text4.grid(row=3, column=5)
outshellexcel = tk.Entry(frame6, width=20, borderwidth=5)
outshellexcel.grid(row=3, column=7)
text4 = tk.Label(frame6, text="C", font=("Helvetica", 10))
text4.grid(row=2, column=8)
text4 = tk.Label(frame6, text="C", font=("Helvetica", 10))
text4.grid(row=3, column=8)
text4 = tk.Label(frame6, text="Specific heat", font=("Helvetica", 10))
text4.grid(row=4, column=5)
spheat2excel = tk.Entry(frame6, width=20, borderwidth=5)
spheat2excel.grid(row=4, column=7)
text4 = tk.Label(frame6, text="j/Kgk", font=("Helvetica", 10))
text4.grid(row=4, column=8)
text4 = tk.Label(frame6, text="Density", font=("Helvetica", 10))
text4.grid(row=5, column=5)
den2excel = tk.Entry(frame6, width=20, borderwidth=5)
den2excel.grid(row=5, column=7)
text4 = tk.Label(frame6, text="Kg/m3", font=("Helvetica", 10))
text4.grid(row=5, column=8)
text4 = tk.Label(frame6, text="Dynamic Viscosity", font=("Helvetica", 10))
text4.grid(row=6, column=5)
vis2excel = tk.Entry(frame6, width=20, borderwidth=5)
vis2excel.grid(row=6, column=7)
text4 = tk.Label(frame6, text="Ns/m2", font=("Helvetica", 10))
text4.grid(row=6, column=8)
text4 = tk.Label(frame6, text="Thermal Conductivity", font=("Helvetica", 10))
text4.grid(row=7, column=5)
tcon2excel = tk.Entry(frame6, width=20, borderwidth=5)
tcon2excel.grid(row=7, column=7)
text4 = tk.Label(frame6, text="W/m.⁰C", font=("Helvetica", 10))
text4.grid(row=7, column=8)
text4 = tk.Label(frame6, text="kg/hr", font=("Helvetica", 10))
text4.grid(row=8, column=8)
text4 = tk.Label(frame6, text="Mass flow rate", font=("Helvetica", 10))
text4.grid(row=8, column=5)
mflow2excel = tk.Entry(frame6, width=20, borderwidth=5)
mflow2excel.grid(row=8, column=7)
text4 = tk.Label(frame6, text="fouling(shell)", font=("Helvetica", 10))
text4.grid(row=9, column=5)
foulsexcel = tk.Entry(frame6, width=20, borderwidth=5)
foulsexcel.grid(row=9, column=7)

# --------------------------Common Excel Part
text4 = tk.Label(frame6, text="Ft ", font=("Helvetica", 10))
text4.grid(row=2, column=10)
fowlingexcel = tk.Entry(frame6, width=20, borderwidth=5)
fowlingexcel.grid(row=2, column=11)


# f2btn = tk.Button(frame6, text="Coefficients", padx=30, pady=10, command=heatfac)
# f2btn.grid(row=12, column=11)

f2btn = tk.Button(frame6, text="Ft value", padx=40, pady=10, command=lambda: show_frame(frame11))
f2btn.grid(row=13, column=11)

# ------------------------------R & S Value
text4 = tk.Label(frame6, text="R", font=("Helvetica", 10))
text4.grid(row=12, column=5)
rvalueexcel = tk.Entry(frame6, width=20, borderwidth=5)
rvalueexcel.grid(row=12, column=7)
text4 = tk.Label(frame6, text="S", font=("Helvetica", 10))
text4.grid(row=12, column=1)
svalueexcel = tk.Entry(frame6, width=20, borderwidth=5)
svalueexcel.grid(row=12, column=3)
f2btn = tk.Button(frame6, text="Value", padx=10, pady=5, command=crop1)
f2btn.grid(row=12, column=8)

# -----------------  Calculate Button

f2btn = tk.Button(frame6, text="Calculate Excel", padx=40, pady=10, command=shellandtubeExcel)
f2btn.grid(row=14, column=5)

# -------------------------------- Default
show_frame(frame0)
window.mainloop()
